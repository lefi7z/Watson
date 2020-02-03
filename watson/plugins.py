import os
import sys
import shutil
import platform

import arrow
from openpyxl import load_workbook
from openpyxl.styles.fonts import Font

# TODO: what for??
print(platform.python_version())


def flushex(watson, filename, month, projects=None, tags=None,
            round_to_quarter=False):
    """Write the contents of the logfile for one month to an Excel sheet."""
    frames = watson.frames
    months = {'jan': 1,
              'feb': 2,
              'mar': 3, 'mär': 3,
              'apr': 4,
              'may': 5, 'mai': 5,
              'jun': 6,
              'jul': 7,
              'aug': 8,
              'sep': 9,
              'oct': 10, 'okt': 10,
              'nov': 11,
              'dec': 12, 'dez': 12,
             }
    try:
        num_month = months[month[:3].lower()]
    except KeyError:
        if month == 'last':
            num_month = arrow.now().shift(months=-1).month
        else:
            raise
    this_year = arrow.now().year
    if num_month == 12:
        this_year -= 1  # in January, usually one wants to update last year
    first_day = arrow.Arrow(year=this_year, month=num_month, day=1)
    span = frames.span(first_day, first_day.shift(months=1))
    filtered_frames = frames.filter(projects, tags, span) 

    print('loading Excel file...')
    wb = load_workbook(filename)
    print('Excel file loaded.')
    ws = wb[(wb.sheetnames[num_month])]
    for i in range(35, 35 + 4 * 23 + 1):  # empty rows
        for alpha in ('D', 'E', 'F'):
            adress = '%s%i' % (alpha, i)
            print('writing to', adress)
            ws[adress].value = ''
    for f in filtered_frames:
        day = f.start.day
        # get number of weekend days:
        walk = first_day
        weekend_days = 0
        while walk.day <= day and walk.month == num_month:
            if walk.weekday() > 4:  # (saturday: 5, sunday: 6)
                weekend_days += 1
            walk = walk.shift(days=1)
        row = 35 + 4*(day-1-weekend_days)
        skip = 0
        while ws['D%d' % (row+skip)].value:
            skip += 1
            if skip > 4:
                raise IOError("not enough free cells!")
        print('writing day %d to row %d...' % (day, row+skip))
        start_cell = ws['D%d' % (row+skip)]
        start_cell.font = Font(size=8)
        start_cell.number_format = 'HH:MM'
        start_cell.value = f.start.time()
        stop_cell = ws['E%d' % (row+skip)]
        stop_cell.font = Font(size=8)
        stop_cell.number_format = 'HH:MM'
        stop_cell.value = f.stop.time()

    print('making backup...')
    bup = filename + '.bup'
    shutil.copy(filename, bup)
    print('saving Excel file...')
    wb.save(filename)

